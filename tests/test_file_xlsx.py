import os

import openpyxl
from jones_lib.models import FileXLSX 

def test_file_xlsx_read():
    current_dir = os.path.dirname(os.path.abspath(__file__))
    sample_xlsx_path = os.path.join(current_dir, 'resources', 'sample.xlsx')

    file_csv = FileXLSX (sample_xlsx_path)

    workbook = openpyxl.load_workbook(sample_xlsx_path)
    sheet = workbook.active
    data = []

    keys = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]

    for row in sheet.iter_rows(min_row=2):
        data.append({keys[i]: cell.value for i, cell in enumerate(row)})

    assert data == file_csv.read()