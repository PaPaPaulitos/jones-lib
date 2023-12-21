from typing import Dict, Any
from ..interfaces import IFile
import openpyxl

class FileXLSX(IFile):
    def __init__(self, file):
        self.file = file

    def read(self) -> Dict[str, Any]:
        workbook = openpyxl.load_workbook(self.file)
        sheet = workbook.active
        data = []

        keys = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]

        for row in sheet.iter_rows(min_row=2):
            data.append({keys[i]: cell.value for i, cell in enumerate(row)})

        return data
